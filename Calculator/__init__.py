import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from datetime import datetime, date
import json
class DeliveryFinancialTable:
    def __init__(self, df: pd.DataFrame, date: date, collector_zp: int, drivers_zp: int, additional_costs: dict,
                 deliveries_total: int, deliveries_internal: int,  deliveries_distance: int, driver_hours: float):
        self.df = df
        self.date = date
        self.collector_zp = collector_zp
        self.drivers_zp = drivers_zp
        self.additional_costs = additional_costs
        self.deliveries_total = deliveries_total
        self.deliveries_internal = deliveries_internal
        self.deliveries_distance = deliveries_distance
        self.driver_hours = driver_hours
        self.weekday = self.date.strftime("%A").lower()
        self.sum_without_discount = df[df["Delivery.IsDelivery"] == "DELIVERY_ORDER"]['DishSumInt'].sum()
        self.fields = {
            "Сумма без скидки": self.sum_without_discount,
            "Сумма скидки (А)": None,
            "A|Стартер": None,
            "A|Остальное": None,
            "Комиссия (B)": None,
            "B|Комиссия сайтов (I)": None,
            "B1|Я.еда|Внешний курьер": None,
            "B1|Я.еда|Наш курьер + Самовывоз": None,
            "B1|Delivery|Наш курьер + Самовывоз": None,
            "B1|Сайт chiko2.ru": None,
            "B|Комиссия банка (II)": None,
            "B2|Комиссия терминалы": None,
            "B2|Интернет эквайринг": None,
            "Сырье (С)": None,
            "C|Амато (Am)": None,
            "C|Амато|Кухня": None,
            "C|Амато|Кондитерка": None,
            "C|Амато|Бар": None,
            "C|Амато|Упаковка": None,
            "C|МиГуста (Mg)": None,
            "C|МиГуста|Кухня": None,
            "C|МиГуста|Кондитерка": None,
            "C|МиГуста|Упаковка": None,
            "C|Шу-Шу (Shu)": None,
            "C|Шу-Шу|Кухня": None,
            "C|Шу-Шу|Кондитерка": None,
            "C|Шу-Шу|Упаковка": None,
            "C|Угли (Ug)": None,
            "C|Угли|Кухня": None,
            "C|Угли|Бар": None,
            "C|Угли|Упаковка": None,
            "Оплата труда (D)": None,
            "D|Производство (Factory)": None,
            "D|Производство|Амато|Кухня": None,
            "D|Производство|Амато|Кондитерка": None,
            "D|Производство|Амато|Бар": None,
            "D|Производство|МиГуста|Кухня": None,
            "D|Производство|МиГуста|Кондитерка": None,
            "D|Производство|Шу-Шу|Кухня": None,
            "D|Производство|Шу-Шу|Кондитерка": None,
            "D|Производство|Угли|Кухня": None,
            "D|Производство|Угли|Бар": None,
            "D|Сервис (Service)": None,
            "D|Сервис|Оператор": None,
            "D|Сервис|Сборщик": None,
            "D|Сервис|Курьеры": None,
            "D|Сервис|Дополнительно": None,
            "D|Управление (Management)": None,
            "D|Управление|Шеф-повара": None,
            "D|Управление|Бухгалтер/Технолог": None,
            "D|Управление|Архитектор": None,
            "D|Управление|Директор/Зам директор": None,
            "D|Управление|СММ": None,
            "D|Управление|Закупщик": None,
            "D|Управление|Отзывы": None,
            "ГСМ (Е)": None,
            "Аренда/Коммуналка (F)": None,
            "Прибыль": None,
            "Стоимость услуги доставки": None,
            "Количество услуг водителей всего": None,
            "Количество услуг внутренних": None,
            "Средняя дистанция доставки": None,
            "Количество часов водителей": None
        }
        self.hierarchy = {
            "B|Комиссия сайтов (I)": ["B1|Я.еда|Внешний курьер", "B1|Я.еда|Наш курьер + Самовывоз", "B1|Delivery|Наш курьер + Самовывоз", "B1|Сайт chiko2.ru"],
            "B|Комиссия банка (II)": ["B2|Комиссия терминалы", "B2|Интернет эквайринг"],
            "C|Амато (Am)": ["C|Амато|Кухня", "C|Амато|Кондитерка", "C|Амато|Бар", "C|Амато|Упаковка"],
            "C|МиГуста (Mg)": ["C|МиГуста|Кухня", "C|МиГуста|Кондитерка", "C|МиГуста|Упаковка"],
            "C|Шу-Шу (Shu)": ["C|Шу-Шу|Кухня", "C|Шу-Шу|Кондитерка", "C|Шу-Шу|Упаковка"],
            "C|Угли (Ug)": ["C|Угли|Кухня", "C|Угли|Бар", "C|Угли|Упаковка"],
            "D|Производство (Factory)": ["D|Производство|Амато|Кухня", "D|Производство|Амато|Кондитерка", "D|Производство|Амато|Бар", "D|Производство|МиГуста|Кухня", "D|Производство|МиГуста|Кондитерка", "D|Производство|Шу-Шу|Кухня", "D|Производство|Шу-Шу|Кондитерка", "D|Производство|Угли|Кухня", "D|Производство|Угли|Бар"],
            "D|Сервис (Service)": ["D|Сервис|Оператор", "D|Сервис|Сборщик", "D|Сервис|Курьеры", "D|Сервис|Дополнительно"],
            "D|Управление (Management)": ["D|Управление|Шеф-повара", "D|Управление|Бухгалтер/Технолог", "D|Управление|Архитектор", "D|Управление|Директор/Зам директор", "D|Управление|СММ", "D|Управление|Закупщик", "D|Управление|Отзывы"],
            "Сумма скидки (А)": ["A|Стартер", "A|Остальное"],
            "Комиссия (B)": ["B|Комиссия сайтов (I)", "B|Комиссия банка (II)"],
            "Сырье (С)": ["C|Амато (Am)", "C|МиГуста (Mg)", "C|Шу-Шу (Shu)", "C|Угли (Ug)"],
            "Оплата труда (D)": ["D|Производство (Factory)", "D|Сервис (Service)", "D|Управление (Management)"]
        }
        # Add data rows
        self.summary_rows = ["Сумма без скидки", "Сумма скидки (А)", "Комиссия (B)", "Сырье (С)", "Оплата труда (D)", "ГСМ (Е)", "Аренда/Коммуналка (F)"]
        self.secondary_summary_rows = ["B|Комиссия сайтов (I)", "B|Комиссия банка (II)", "C|Амато (Am)", "C|МиГуста (Mg)", "C|Шу-Шу (Shu)", "C|Угли (Ug)", "D|Производство (Factory)", "D|Сервис (Service)", "D|Управление (Management)"]

    def calculate_totals(self):
        for key, subfields in self.hierarchy.items():
            #print(key, subfields)
            self.fields[key] = sum(self.fields[subfield] for subfield in subfields if subfield in self.fields)

    def calculate_percentage(self, field_value):
        if self.sum_without_discount > 0:
            return (field_value / self.sum_without_discount) * 100
        return 0

    def display(self):
        for field, value in self.fields.items():
            if field not in ["Сумма без скидки", "Прибыль"]:
                self.fields[field] = self.get_field(field) if field not in self.hierarchy.keys() else value
        self.calculate_totals()
        self.fields["Прибыль"] = self.get_field("Прибыль")
        print(f"{'Название':<40}{'Сумма':>9}{'Процент':>6}")
        for field, value in self.fields.items():
            percentage = self.calculate_percentage(value)
            print(f"{field:<40}{value:>9.2f}{percentage:>6.2f}")

    def save_to_excel(self, file_path):
        for field, value in self.fields.items():
            if field not in ["Сумма без скидки", "Прибыль"]:
                self.fields[field] = self.get_field(field) if field not in self.hierarchy.keys() else value
        for field, value in self.fields.items():
            self.fields[field] = value if value != None else 0
        self.calculate_totals()
        self.fields["Прибыль"] = self.get_field("Прибыль")
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Table"

        header_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        summary_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        secondary_summary_fill = PatternFill(start_color="FFDD88", end_color="FFDD88", fill_type="solid")
        total_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
        profit_fill = PatternFill(start_color="CCFF99", end_color="CCFF99", fill_type="solid")
        bold_border = Border(left=Side(style="thin"),
                             right=Side(style="thin"),
                             top=Side(style="thin"),
                             bottom=Side(style="thin"))

        # Add header row
        ws.append(["Название", "Сумма", "Процент"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.border = bold_border

        
        for field, value in self.fields.items():
            percentage = self.calculate_percentage(self.fields[field])
            row = ws.max_row + 1
            ws.append([field, round(value, 2) , round(percentage, 2)])

            # Highlight specific rows
            if field == "Сумма без скидки":
                for cell in ws[row]:
                    cell.fill = total_fill
                    cell.border = bold_border
            elif field == "Прибыль":
                for cell in ws[row]:
                    cell.fill = profit_fill
                    cell.border = bold_border
            elif field in self.summary_rows:
                for cell in ws[row]:
                    cell.fill = summary_fill
                    cell.border = bold_border
            elif field in self.secondary_summary_rows:
                for cell in ws[row]:
                    cell.fill = secondary_summary_fill
                    cell.border = bold_border
            else:
                for cell in ws[row]:
                    cell.border = bold_border

        # Add bold border for the first column
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.border = bold_border
        file_path = f"Saved/{file_path}"
        wb.save(file_path)
        print(f"Excel file saved at: {file_path}")
        return file_path

    def get_field(self, field_name):
        data = None
        with open("info.json", "r", encoding="utf-8") as file:
            data = json.load(file)

        ya_ext_commission = data["ya_ext_commission"]
        ya_ext_order_cost = data["ya_ext_order_cost"]
        ya_int_commission = data["ya_int_commission"]
        delivery_comission = data["delivery_comission"]
        starter_comission = data["starter_comission"]
        terminal_comission = data["terminal_comission"]
        internet_comission = data["internet_comission"]
        amato_factory_mon_thurs = data["amato_factory_mon_thurs"]
        amato_factory_fri_sun = data["amato_factory_fri_sun"]
        shu_confetioner = data["shu_confetioner"]
        amato_bar = data["amato_bar"]
        migusta_factory_sun_thurs = data["migusta_factory_sun_thurs"]
        migusta_factory_fri_sat = data["migusta_factory_fri_sat"]
        shu_factory = data["shu_factory"]
        ugli_factory = data["ugli_factory"]
        ugli_bar = data["ugli_bar"]
        operator_comission = data["operator_comission"]
        operator_salary = data["operator_salary"]
        zp_maks_chief = data["zp_maks_chief"]
        zp_slava_chief = data["zp_slava_chief"]
        zp_architect = data["zp_architect"]
        zp_director = data["zp_director"]
        zp_zamdir = data["zp_zamdir"]
        zp_smm = data["zp_smm"]
        zp_buyer = data["zp_buyer"]
        zp_reviews = data["zp_reviews"]
        zp_buh_teh = data["zp_buh_teh"]
        payment_per_km = data["payment_per_km"]
        arenda_kommunalka = data["arenda_kommunalka"]
        if field_name not in self.fields:
            raise ValueError(f"Field '{field_name}' not found in the financial table.")
        else:
            d_df = self.df.loc[self.df["Delivery.IsDelivery"] == "DELIVERY_ORDER"]
            if field_name == "A|Стартер":
                return d_df.loc[d_df["ItemSaleEventDiscountType"] == "Стартер"]['DiscountSum'].sum()
            elif field_name == "A|Остальное":
                return d_df.loc[d_df["ItemSaleEventDiscountType"] != "Стартер"]['DiscountSum'].sum()
            elif field_name == "B1|Я.еда|Внешний курьер":
                temp_df = d_df.loc[(d_df["OrderType"] == "Яндекс|Внешний курьер") & 
                                (d_df["OriginName"] == "yandex_food")]
                print(temp_df["DishDiscountSumInt"].sum(), ya_ext_commission, temp_df["UniqOrderId.OrdersCount"].sum(), ya_ext_order_cost)
                return temp_df["DishDiscountSumInt"].sum() * ya_ext_commission + temp_df["UniqOrderId.OrdersCount"].sum() * ya_ext_order_cost
            elif field_name == "B1|Я.еда|Наш курьер + Самовывоз":
                return d_df.loc[d_df["OrderType"].isin(["Яндекс|Наш курьер", "Яндекс|Самовывоз гостем"]) & 
                                (d_df["OriginName"] == "yandex_food")]["DishDiscountSumInt"].sum() * ya_int_commission
            elif field_name == "B1|Delivery|Наш курьер + Самовывоз":
                return d_df.loc[d_df["OrderType"].isin(["Яндекс|Наш курьер", "Яндекс|Самовывоз гостем"]) & 
                                (d_df["OriginName"] == "delivery_club")]["DishDiscountSumInt"].sum() * delivery_comission
            elif field_name == "B1|Сайт chiko2.ru":
                return d_df.loc[d_df["OriginName"] == "Starter"]["DishDiscountSumInt"].sum() * starter_comission
            elif field_name == "B2|Комиссия терминалы":
                return d_df.loc[d_df["PayTypes"] == "Банковские карты"]["DishDiscountSumInt"].sum() * terminal_comission    
            elif field_name == "B2|Интернет эквайринг":
                return d_df.loc[d_df["PayTypes"] == "Оплата онлайн Starter"]["DishDiscountSumInt"].sum() * internet_comission
            elif field_name == "C|Амато|Кухня":
                #print(d_df[["DishCategory", "RestorauntGroup", "ProductCostBase.ProductCost", "Delivery.IsDelivery"]])
                return d_df.loc[(d_df["DishCategory"].isin(["кухня", "кухня (исключение)"])) &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Амато|Кондитерка":
                return d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Амато|Бар":
                return d_df.loc[(d_df["DishCategory"] == "бар") &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Амато|Упаковка":
                return d_df.loc[(d_df["DishCategory"] == "упаковка бесплатная") &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|МиГуста|Кухня":
                return d_df.loc[(d_df["DishCategory"] == "кухня") &
                                (d_df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|МиГуста|Кондитерка":
                return d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|МиГуста|Упаковка":
                return d_df.loc[(d_df["DishCategory"] == "упаковка бесплатная") &
                                (d_df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Шу-Шу|Кухня":
                return d_df.loc[(d_df["DishCategory"] == "кухня") &
                                (d_df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Шу-Шу|Кондитерка":
                return d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Шу-Шу|Упаковка":
                return d_df.loc[(d_df["DishCategory"] == "упаковка бесплатная") &
                                (d_df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Угли|Кухня":
                return d_df.loc[(d_df["DishCategory"] == "кухня") &
                                (d_df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Угли|Бар":
                return d_df.loc[(d_df["DishCategory"] == "бар") &
                                (d_df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "C|Угли|Упаковка":
                return d_df.loc[(d_df["DishCategory"] == "упаковка бесплатная") &
                                (d_df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum()
            elif field_name == "D|Производство|Амато|Кухня":
                temp_val = amato_factory_mon_thurs if self.weekday in ["monday", "tuesday", "wednesday", "thursday"] else amato_factory_fri_sun
                # 25к и 30к
                temp_k = d_df.loc[(d_df["DishCategory"].isin(["кухня", "кухня (исключение)"])) &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"] == "кухня") &
                                                                                                                            (self.df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Амато|Кондитерка":
                temp_val = shu_confetioner
                # 13к в день
                temp_k = d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum() / self.df.loc[self.df["DishCategory"] == "кондитерка"]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Амато|Бар":
                temp_val = amato_bar
                # 5к
                temp_k = d_df.loc[(d_df["DishCategory"] == "бар") &
                                (d_df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"] == "бар") &
                                                                                                                            (self.df["RestorauntGroup"] == "Амато")]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|МиГуста|Кухня":
                temp_val = migusta_factory_sun_thurs if self.weekday in ["monday", "tuesday", "wednesday", "thursday", "sunday"] else migusta_factory_fri_sat
                # 23250 и 25125
                temp_k = d_df.loc[(d_df["DishCategory"].isin(["кухня", "кухня (исключение)"])) &
                                (d_df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"].isin(["кухня", "кухня (исключение)"])) &
                                                                                                                  (self.df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum()
                #print(temp_k, temp_val)
                return temp_val * temp_k
            elif field_name == "D|Производство|МиГуста|Кондитерка":
                temp_val = shu_confetioner
                # 13к в день
                temp_k = d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "МиГуста")]["ProductCostBase.ProductCost"].sum() / self.df.loc[self.df["DishCategory"] == "кондитерка"]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Шу-Шу|Кухня":
                temp_val = shu_factory
                # 7500 в день
                temp_k = d_df.loc[(d_df["DishCategory"] == "кухня") &
                                (d_df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"] == "кухня") &
                                                                                                                            (self.df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Шу-Шу|Кондитерка":
                temp_val = shu_confetioner
                # 13к в день
                temp_k = d_df.loc[(d_df["DishCategory"] == "кондитерка") &
                                (d_df["RestorauntGroup"] == "Шу-Шу")]["ProductCostBase.ProductCost"].sum() / self.df.loc[self.df["DishCategory"] == "кондитерка"]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Угли|Кухня":
                temp_val = ugli_factory
                # 4500 в день
                temp_k = d_df.loc[(d_df["DishCategory"] == "кухня") &
                                (d_df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"] == "кухня") &
                                                                                                                            (self.df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k
            elif field_name == "D|Производство|Угли|Бар":
                # 1500 в день
                temp_val = ugli_bar
                temp_k = d_df.loc[(d_df["DishCategory"] == "бар") &
                                (d_df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum() / self.df.loc[(self.df["DishCategory"] == "бар") &
                                                                                                                            (self.df["RestorauntGroup"] == "Гриль бар")]["ProductCostBase.ProductCost"].sum()
                return temp_val * temp_k

            elif field_name == "D|Сервис|Оператор":
                return d_df["DishDiscountSumInt"].sum() * operator_comission + operator_salary
            elif field_name == "D|Сервис|Сборщик":
                return self.collector_zp
            elif field_name == "D|Сервис|Курьеры":
                return self.drivers_zp
            elif field_name == "D|Сервис|Дополнительно":
                return self.additional_costs
            elif field_name == "D|Управление|Шеф-повара":
                temp_df_general = self.df.loc[self.df["DishCategory"].isin(["кухня", "кондитерка", "кухня (исключение)", "слойки"])]
                temp_df_slava = temp_df_general[temp_df_general["RestorauntGroup"].isin(["МиГуста","Гриль бар"]) &
                                                temp_df_general["DishCategory"].isin(["кухня"])]
                #Макс - это все из general, кроме Слава
                temp_df_maks = temp_df_general[~temp_df_general.index.isin(temp_df_slava.index)]
                maks = zp_maks_chief/30 * temp_df_maks[temp_df_maks["Delivery.IsDelivery"] == "DELIVERY_ORDER"]['DishSumInt'].sum()/temp_df_maks["DishSumInt"].sum() 
                slava = zp_slava_chief/30 * temp_df_slava[temp_df_slava["Delivery.IsDelivery"] == "DELIVERY_ORDER"]['DishSumInt'].sum()/temp_df_slava["DishSumInt"].sum()
                return maks + slava
            elif field_name == "D|Управление|Бухгалтер/Технолог":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * zp_buh_teh/30
            elif field_name == "D|Управление|Архитектор":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * zp_architect/30
            elif field_name == "D|Управление|Директор/Зам директор":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * (zp_director+zp_zamdir)/30
            elif field_name == "D|Управление|СММ":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * zp_smm/30
            elif field_name == "D|Управление|Закупщик":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * zp_buyer/30
            elif field_name == "D|Управление|Отзывы":
                return self.df.loc[self.df['Delivery.IsDelivery'] == 'DELIVERY_ORDER']['DishSumInt'].sum()/self.df['DishSumInt'].sum() * zp_reviews/30
            elif field_name == "ГСМ (Е)":
                return self.deliveries_distance * payment_per_km
            elif field_name == "Аренда/Коммуналка (F)":
                return arenda_kommunalka/30
            elif field_name == "Количество услуг водителей всего":
                return self.deliveries_total
            elif field_name == "Количество услуг внутренних":
                return self.deliveries_internal
            elif field_name == "Средняя дистанция доставки":
                return self.deliveries_distance/self.deliveries_total
            elif field_name == "Стоимость услуги доставки":
                return (self.drivers_zp + self.additional_costs)/(self.deliveries_total)
            elif field_name == "Количество часов водителей":
                return self.driver_hours
            elif field_name == "Прибыль":
                #print(self.fields["Сумма без скидки"],
                    #   self.fields['Сумма скидки (А)'],
                    #   self.fields["Комиссия (B)"],
                    #   self.fields["Сырье (С)"],
                    #   self.fields["Оплата труда (D)"],
                    #   self.fields["ГСМ (Е)"],
                    #   self.fields["Аренда/Коммуналка (F)"])
                return self.fields["Сумма без скидки"] - self.fields['Сумма скидки (А)'] - self.fields["Комиссия (B)"] - self.fields["Сырье (С)"] - self.fields["Оплата труда (D)"] - self.fields["ГСМ (Е)"] - self.fields["Аренда/Коммуналка (F)"]
            else:
                return ValueError(f"Field '{field_name}' not found in the financial table.")